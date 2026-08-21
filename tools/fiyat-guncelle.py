#!/usr/bin/env python3
"""evprices.js içindeki FUEL_HIST tablosunu açık kaynaklardan yeniden üretir.

Neden derleme zamanında? Uygulama çevrimdışı çalışıyor ve sunucusu yok:
tarayıcıdan doğrudan çağrı CORS'a takılır, API anahtarı da istemcide görünür
olurdu. Veri burada toplanıp tabloya gömülüyor; uygulama çalışırken ağa hiç
çıkmıyor.

KAYNAKLAR
  AB (27 ülke) — Avrupa Komisyonu Haftalık Petrol Bülteni, "Prices with taxes"
                 sayfası: vergiler dahil tüketici fiyatı, haftalık, 1000 l/EUR.
  Türkiye      — TPPD geçmiş fiyat arşivi, Ankara/Çankaya (aşağıdaki sabit).

ÇIKTI  evprices.js -> FUEL_HIST = {ÜLKE: {src, tur:[petrol,diesel,lpg], m:{'YYYY-MM':[...]}}}
       Değerler ÜLKENİN KENDİ PARA BİRİMİNDE, litre başına, vergiler dahil.
       Ay anahtarı o ayın ortalamasıdır.

Kullanım:  python tools/fiyat-guncelle.py [--yaz]
           --yaz verilmezse hiçbir dosya değiştirilmez, yalnız rapor basılır.
"""
import io, json, re, sys, urllib.request, urllib.error
from collections import defaultdict
from datetime import datetime, timezone

YAZ = '--yaz' in sys.argv
UA = {'User-Agent': 'WattTrack-fiyat-guncelleyici/1.0 (+https://github.com/Rino-06/WattTrack)',
      'Accept-Language': 'tr-TR,tr;q=0.9,en;q=0.8'}

# Ankara / Çankaya — TPPD arşivinde il ve ilçe kodu.
# `county` parametresinin GERÇEKTEN dikkate alındığı sınandı: 1000 ile 1001
# farklı fiyat serisi veriyor, parametresiz ve geçersiz değerde sunucu hata
# döndürüyor. İlçe ADI sayfanın hiçbir yerinde yazmadığı için 1000'in Çankaya
# olduğu koddan doğrulanamıyor; değer değiştirilecekse tarayıcıda açılıp
# gözle teyit edilmeli.
TPPD_IL, TPPD_ILCE = 6, 1000
TPPD_BASLANGIC = '01.01.2020'

def indir(url, limit=90_000_000, ek=None):
    b = dict(UA)
    if ek: b.update(ek)
    r = urllib.request.Request(url, headers=b)
    with urllib.request.urlopen(r, timeout=120) as f:
        return f.read(limit)

def duz(h):
    h = re.sub(r'(?is)<(script|style)[^>]*>.*?</\1>', ' ', h)
    return re.sub(r'\s+', ' ', re.sub(r'(?s)<[^>]+>', ' ', h)).strip()

def sayi(s):
    """'44,45' -> 44.45 · '-' / boş -> None"""
    if s is None: return None
    s = str(s).strip().replace('\xa0', '')
    if not s or s in ('-', '—', '–'): return None
    s = s.replace('.', '').replace(',', '.') if re.match(r'^\d{1,3}(\.\d{3})*(,\d+)?$', s) else s.replace(',', '.')
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None

# ============================================================
# AB — Haftalık Petrol Bülteni
# ============================================================
AB_SAYFA = "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
# Uygulamanın COUNTRIES tablosundaki para birimleri (avro olmayanlar)
AVRO_DISI = {'BG': 'BGN', 'CZ': 'CZK', 'DK': 'DKK', 'HU': 'HUF',
             'PL': 'PLN', 'RO': 'RON', 'SE': 'SEK'}

def ab_dosya_adresi():
    html = indir(AB_SAYFA, 8_000_000).decode('utf-8', 'replace')
    for m in re.finditer(r'href="([^"]*document/download/[^"]*?)"', html, re.I):
        h = m.group(1)
        if h.startswith('/'): h = 'https://energy.ec.europa.eu' + h
        if re.search(r'Prices?_History', h, re.I):
            return h.replace('&amp;', '&')
    raise SystemExit("AB: geçmiş dosyasının bağlantısı bulunamadı")

def ab_cek():
    import openpyxl
    url = ab_dosya_adresi()
    print(f"  AB dosyası: {url[:110]}")
    wb = openpyxl.load_workbook(io.BytesIO(indir(url)), read_only=True, data_only=True)
    ad = next((s for s in wb.sheetnames if 'with tax' in s.lower()), None)
    if not ad: raise SystemExit(f"AB: 'Prices with taxes' sayfası yok — {wb.sheetnames}")
    ws = wb[ad]
    satir = [r for r in ws.iter_rows(values_only=True)]
    baslik, altbaslik = satir[0], satir[1]

    # Blok sınırlarını CTR sütunlarından çıkar; genişlik ülkeden ülkeye DEĞİŞİYOR
    ctr = [i for i, v in enumerate(baslik) if v == 'CTR']
    bloklar = []
    for n, i in enumerate(ctr):
        son = ctr[n + 1] if n + 1 < len(ctr) else len(baslik)
        bloklar.append((i, son))

    def sutun_bul(bas, son, kalip, dislama=None):
        for j in range(bas + 1, son):
            t = str(altbaslik[j] or '')
            if re.search(kalip, t, re.I) and not (dislama and re.search(dislama, t, re.I)):
                return j
        return None

    # ülke kodu ilk veri satırından okunuyor ('AT_' gibi)
    veri = satir[3:]
    ulke_sutun = {}
    for bas, son in bloklar:
        kod = str(veri[0][bas] or '').rstrip('_')
        if len(kod) != 2: continue          # EU_ ortalaması ve boş bloklar atlanır
        ulke_sutun[kod] = {
            'petrol': sutun_bul(bas, son, r'Euro[- ]?super\s*95'),
            'diesel': sutun_bul(bas, son, r'Gas oil automobile', r'chauffage'),
            'lpg':    sutun_bul(bas, son, r'GPL|LPG'),
        }
    print(f"  AB: {len(ulke_sutun)} ülke bloğu çözümlendi")

    # haftalık satırları aya topla (1000 l -> l)
    aylik = defaultdict(lambda: defaultdict(list))
    for r in veri:
        t = r[0]
        if not isinstance(t, datetime): continue
        ay = t.strftime('%Y-%m')
        for kod, sut in ulke_sutun.items():
            for tur, j in sut.items():
                if j is None or j >= len(r): continue
                v = r[j]
                if isinstance(v, (int, float)) and v > 0:
                    aylik[kod][(ay, tur)].append(v / 1000.0)
    return aylik

def fx_serisi(para_birimleri, bas, bit):
    """ECB kurları (frankfurter, anahtarsız). {tarih: {PARA: kur}}"""
    if not para_birimleri: return {}
    url = (f"https://api.frankfurter.dev/v1/{bas}..{bit}"
           f"?base=EUR&symbols={','.join(sorted(para_birimleri))}")
    try:
        d = json.loads(indir(url, 8_000_000))
        return d.get('rates', {})
    except Exception as e:
        print(f"  UYARI: kur serisi alınamadı ({type(e).__name__}) — "
              f"avro olmayan ülkeler ATLANACAK")
        return {}

def ab_tablo():
    aylik = ab_cek()
    aylar = sorted({ay for k in aylik for ay, _ in aylik[k]})
    if not aylar: raise SystemExit("AB: hiç ay üretilmedi")
    kur = fx_serisi(set(AVRO_DISI.values()), aylar[0] + '-01', aylar[-1] + '-28')
    # ay -> ortalama kur
    ay_kur = defaultdict(lambda: defaultdict(list))
    for gun, m in kur.items():
        for p, v in m.items():
            ay_kur[gun[:7]][p].append(v)

    out = {}
    for kod, veri in sorted(aylik.items()):
        para = AVRO_DISI.get(kod)
        m = {}
        for ay in sorted({a for a, _ in veri}):
            katsayi = 1.0
            if para:
                k = ay_kur.get(ay, {}).get(para)
                if not k:
                    continue        # kuru olmayan ayı UYDURMA, atla
                katsayi = sum(k) / len(k)
            sira = []
            for tur in ('petrol', 'diesel', 'lpg'):
                d = veri.get((ay, tur))
                sira.append(round(sum(d) / len(d) * katsayi, 3) if d else None)
            if any(v is not None for v in sira):
                m[ay] = sira
        if m:
            out[kod] = {'src': 'ecoil', 'tur': ['petrol', 'diesel', 'lpg'], 'm': m}
    return out

# ============================================================
# TÜRKİYE — TPPD arşivi (Ankara / Çankaya)
# ============================================================
TR_AY = {'ocak':1,'şubat':2,'mart':3,'nisan':4,'mayıs':5,'haziran':6,
         'temmuz':7,'ağustos':8,'eylül':9,'ekim':10,'kasım':11,'aralık':12}

def tr_tarih(s):
    p = duz(s).lower().split()
    if len(p) != 3: return None
    try:
        return f"{int(p[2]):04d}-{TR_AY[p[1]]:02d}"
    except (KeyError, ValueError):
        return None

def tr_tablo():
    bugun = datetime.now(timezone.utc).strftime('%d.%m.%Y')
    url = (f"https://www.tppd.com.tr/gecmis-akaryakit-fiyatlari"
           f"?id={TPPD_IL}&county={TPPD_ILCE}"
           f"&StartDate={TPPD_BASLANGIC}&EndDate={bugun}")
    print(f"  TR kaynağı: {url}")
    html = indir(url, 12_000_000).decode('utf-8', 'replace')
    tablolar = re.findall(r'(?is)<table[^>]*>.*?</table>', html)
    if not tablolar: raise SystemExit("TR: sayfada tablo yok — düzen değişmiş olabilir")
    satirlar = re.findall(r'(?is)<tr[^>]*>(.*?)</tr>', tablolar[0])
    basliklar = [duz(x) for x in re.findall(r'(?is)<t[dh][^>]*>(.*?)</t[dh]>', satirlar[0])]
    # Sütunları BAŞLIK METNİNDEN eşle — sıraya güvenme
    idx = {}
    for i, b in enumerate(basliklar):
        u = b.upper()
        if 'BENZİN' in u and 'petrol' not in idx: idx['petrol'] = i
        elif 'MOTORİN' in u and 'diesel' not in idx: idx['diesel'] = i
        elif u.startswith('GAZ') and 'YAĞ' not in u and 'lpg' not in idx: idx['lpg'] = i
    print(f"  TR sütun eşlemesi: {idx}  (başlıklar: {basliklar})")
    if 'petrol' not in idx or 'diesel' not in idx:
        raise SystemExit("TR: benzin/motorin sütunu bulunamadı")

    aylik = defaultdict(lambda: defaultdict(list))
    for s in satirlar[1:]:
        h = [duz(x) for x in re.findall(r'(?is)<t[dh][^>]*>(.*?)</t[dh]>', s)]
        if len(h) < 2: continue
        ay = tr_tarih(h[0])
        if not ay: continue
        for tur, i in idx.items():
            if i < len(h):
                v = sayi(h[i])
                if v: aylik[ay][tur].append(v)
    m = {}
    for ay in sorted(aylik):
        sira = []
        for tur in ('petrol', 'diesel', 'lpg'):
            d = aylik[ay].get(tur)
            sira.append(round(sum(d) / len(d), 3) if d else None)
        if any(v is not None for v in sira):
            m[ay] = sira
    return {'src': 'tppd', 'tur': ['petrol', 'diesel', 'lpg'], 'm': m}

# ============================================================
# Akıl sağlığı denetimi — saçma veriyi ASLA yazma
# ============================================================
def denetle(tablo):
    hata = []
    for kod, g in tablo.items():
        if len(g['m']) < 6:
            hata.append(f"{kod}: yalnız {len(g['m'])} ay")
        for ay, v in g['m'].items():
            for tur, x in zip(g['tur'], v):
                if x is None: continue
                if not (0.05 < x < 500):
                    hata.append(f"{kod} {ay} {tur}: {x} makul aralık dışında")
    return hata

def js_yaz(tablo, kaynaklar):
    sat = ["const FUEL_HIST = {"]
    for i, (kod, g) in enumerate(sorted(tablo.items())):
        sat.append(f"  {kod}: {{src: '{g['src']}', tur: ['petrol', 'diesel', 'lpg'], m: {{")
        aylar = sorted(g['m'])
        for j, ay in enumerate(aylar):
            v = ', '.join('null' if x is None else f"{x:g}" for x in g['m'][ay])
            son = '' if j == len(aylar) - 1 else ','
            sat.append(f"    '{ay}': [{v}]{son}")
        sat.append("  }}" + ('' if i == len(tablo) - 1 else ','))
    sat.append("};")
    sat.append("")
    sat.append("const FUEL_HIST_SRC = {")
    for i, (k, v) in enumerate(sorted(kaynaklar.items())):
        son = '' if i == len(kaynaklar) - 1 else ','
        sat.append(f"  {k}: {{ad: '{v['ad']}',")
        sat.append(f"    url: '{v['url']}',")
        sat.append(f"    guncel: '{v['guncel']}'}}{son}")
    sat.append("};")
    return "\n".join(sat)

def main():
    print("=" * 70); print("AB — Haftalık Petrol Bülteni"); print("=" * 70)
    tablo = ab_tablo()
    print("=" * 70); print("TÜRKİYE — TPPD arşivi (Ankara / Çankaya)"); print("=" * 70)
    try:
        tablo['TR'] = tr_tablo()
    except SystemExit as e:
        print(f"  TR ATLANDI: {e}")

    print()
    print("=" * 70); print("ÖZET"); print("=" * 70)
    for kod, g in sorted(tablo.items()):
        aylar = sorted(g['m'])
        ilk, son = aylar[0], aylar[-1]
        ornek = g['m'][son]
        print(f"  {kod}: {len(aylar):>3} ay  {ilk} … {son}   son değer={ornek}")

    hatalar = denetle(tablo)
    if hatalar:
        print("\nDENETİM UYARILARI:")
        for h in hatalar[:20]: print("   -", h)
        if len(hatalar) > 20: print(f"   … {len(hatalar)-20} uyarı daha")

    kaynaklar = {
        'ecoil': {'ad': 'Avrupa Komisyonu Haftalık Petrol Bülteni',
                  'url': AB_SAYFA,
                  'guncel': datetime.now(timezone.utc).strftime('%Y-%m-%d')},
        'tppd': {'ad': 'TPPD Geçmiş Akaryakıt Fiyatları (Ankara/Çankaya)',
                 'url': 'https://www.tppd.com.tr/gecmis-akaryakit-fiyatlari',
                 'guncel': datetime.now(timezone.utc).strftime('%Y-%m-%d')},
    }
    kaynaklar = {k: v for k, v in kaynaklar.items()
                 if any(g['src'] == k for g in tablo.values())}
    blok = js_yaz(tablo, kaynaklar)
    print(f"\nüretilen JS: {len(blok.splitlines())} satır")

    if not YAZ:
        print("\n(--yaz verilmedi: dosya DEĞİŞTİRİLMEDİ)")
        print("\n--- ilk 12 satır ---")
        for l in blok.splitlines()[:12]: print("  ", l)
        return

    if hatalar:
        raise SystemExit("Denetim uyarısı varken yazılmıyor.")
    kaynak = open('evprices.js', encoding='utf-8').read()
    yeni = re.sub(r'const FUEL_HIST = \{.*?\n\};\n\nconst FUEL_HIST_SRC = \{.*?\n\};',
                  blok, kaynak, flags=re.S)
    if yeni == kaynak:
        raise SystemExit("evprices.js içinde FUEL_HIST bloğu bulunamadı")
    open('evprices.js', 'w', encoding='utf-8').write(yeni)
    print("evprices.js güncellendi.")

if __name__ == '__main__':
    main()
