"""AB Haftalık Petrol Bülteni sondası.

Sorular:
  1. Veri dosyası hangi adresten, hangi biçimde iniyor?
  2. Hangi ülkeleri kapsıyor — TÜRKİYE var mı?
  3. Hangi yakıt türleri ve hangi tarih aralığı var?
  4. Vergiler dahil tüketici fiyatı mı (uygulamanın ihtiyacı bu)?
"""
import io, re, sys, json
import urllib.request

UA = {'User-Agent': 'Mozilla/5.0 (compatible; fuel-data-probe)'}

def al(url, limit=40_000_000):
    try:
        r = urllib.request.Request(url, headers=UA)
        with urllib.request.urlopen(r, timeout=60) as f:
            return f.status, f.headers.get('Content-Type', ''), f.read(limit)
    except Exception as e:
        return None, str(e), b''

print("=" * 70)
print("ADIM 1 — Bülten sayfasını bul, veri dosyası bağlantılarını çıkar")
print("=" * 70)

landing = [
    "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en",
    "https://energy.ec.europa.eu/document/download/",
]
links = []
for u in landing[:1]:
    st, ct, body = al(u)
    print(f"\n{u}\n  durum={st}  tip={ct[:60]}  boyut={len(body)}")
    if st == 200:
        html = body.decode('utf-8', 'replace')
        for m in re.finditer(r'href="([^"]+\.(?:xlsx|xls|csv|ods)[^"]*)"', html, re.I):
            h = m.group(1)
            if h.startswith('/'):
                h = 'https://energy.ec.europa.eu' + h
            links.append(h)
        for m in re.finditer(r'href="([^"]*document/download/[^"]+)"', html, re.I):
            h = m.group(1)
            if h.startswith('/'):
                h = 'https://energy.ec.europa.eu' + h
            links.append(h)
        print(f"  bulunan veri bağlantısı: {len(links)}")
        for h in dict.fromkeys(links[:12]):
            print("   -", h[:150])

print()
print("=" * 70)
print("ADIM 2 — AB Açık Veri Portalı'nda veri seti ara")
print("=" * 70)
odp = "https://data.europa.eu/api/hub/search/search?q=weekly%20oil%20bulletin&limit=5"
st, ct, body = al(odp)
print(f"{odp[:80]}\n  durum={st}  tip={ct[:60]}")
if st == 200:
    try:
        d = json.loads(body)
        res = d.get('result', {}).get('results', [])
        print(f"  sonuç: {len(res)}")
        for r in res[:5]:
            t = r.get('title', {})
            t = t.get('en') if isinstance(t, dict) else t
            print("   -", str(t)[:100])
            for dist in (r.get('distributions') or [])[:4]:
                url = dist.get('access_url') or dist.get('download_url')
                if isinstance(url, list): url = url[0] if url else None
                fmt = dist.get('format', {})
                fmt = fmt.get('label') if isinstance(fmt, dict) else fmt
                if url:
                    print(f"       [{fmt}] {str(url)[:130]}")
                    if re.search(r'\.(xlsx|xls|csv|ods)$', str(url), re.I):
                        links.append(str(url))
    except Exception as e:
        print("  ayrıştırılamadı:", e)

print()
print("=" * 70)
print("ADIM 3 — Bilinen doğrudan adresleri dene")
print("=" * 70)
bilinen = [
    "https://ec.europa.eu/energy/observatory/reports/Oil_Bulletin_Prices_History.xlsx",
    "https://energy.ec.europa.eu/system/files/Oil_Bulletin_Prices_History.xlsx",
    "https://energy.ec.europa.eu/document/download/Oil_Bulletin_Prices_History.xlsx",
]
for u in bilinen:
    st, ct, body = al(u, 2_000_000)
    print(f"  durum={st!s:6} tip={ct[:40]:40} boyut={len(body):>9}  {u[:90]}")
    if st == 200 and len(body) > 10000:
        links.append(u)

links = list(dict.fromkeys(links))
print(f"\nToplam aday veri dosyası: {len(links)}")

print()
print("=" * 70)
print("ADIM 4 — İlk çalışan veri dosyasını AÇ ve içeriğini incele")
print("=" * 70)
try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("openpyxl yok")

incelendi = False
for u in links:
    if incelendi: break
    st, ct, body = al(u, 60_000_000)
    if st != 200 or len(body) < 10000:
        continue
    print(f"\nİNDİRİLDİ: {u[:130]}")
    print(f"  tip={ct[:60]}  boyut={len(body)}")
    if u.lower().endswith(('.xlsx', '.xls', '.ods')) or 'sheet' in ct or 'excel' in ct:
        if not openpyxl:
            continue
        try:
            wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        except Exception as e:
            print("  açılamadı:", e); continue
        print("  sayfalar:", wb.sheetnames[:12])
        ws = wb[wb.sheetnames[0]]
        satirlar = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            satirlar.append(row)
            if i > 60: break
        print("\n  --- ilk 12 satır ---")
        for r in satirlar[:12]:
            hucre = [str(c)[:22] for c in r[:10] if c is not None]
            if hucre: print("   |", " | ".join(hucre))
        metin = " ".join(str(c) for r in satirlar for c in r if c is not None)
        print("\n  --- TÜRKİYE ARAMASI ---")
        for kalip in ['Turkey', 'Türkiye', 'Turkiye', 'TR ', '(TR)']:
            print(f"   {kalip!r}: {'VAR' if kalip in metin else 'yok'}")
        print("\n  --- yakıt türü ipuçları ---")
        for kalip in ['Euro-super 95', 'Euro Super 95', 'Gas oil', 'Diesel',
                      'LPG', 'Heating', 'taxes', 'Taxes', 'VAT']:
            if kalip in metin: print("   VAR:", kalip)
        incelendi = True
    elif u.lower().endswith('.csv') or 'csv' in ct:
        txt = body.decode('utf-8', 'replace')
        print("  --- ilk 6 satır ---")
        for l in txt.split('\n')[:6]: print("   |", l[:160])
        print("  Türkiye:", 'VAR' if re.search(r'Turkey|Türkiye', txt) else 'yok')
        incelendi = True

if not incelendi:
    print("\nHiçbir veri dosyası açılamadı — ADIM 1/2 çıktısına bakılmalı.")
print("\n" + "=" * 70)
print("SONDA BİTTİ")
print("=" * 70)
